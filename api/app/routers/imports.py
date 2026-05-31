import csv
import io
from typing import Iterable

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import Classe, Ecue, Enseignant, Niveau, Semestre
from app.schemas import ImportResult
from app.security import require_admin

router = APIRouter(prefix="/import", tags=["imports"])


# ─── Lecture polymorphique CSV / XLSX ─────────────────────────────────────────


def _read_csv(content: bytes) -> list[dict]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    return [{(k or "").strip(): (v or "").strip() for k, v in row.items()} for row in reader]


def _read_xlsx(content: bytes) -> list[dict]:
    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return []
    rows: list[dict] = []
    headers: list[str] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        cells = [("" if v is None else str(v)).strip() for v in row]
        if i == 0:
            headers = [h.lower() for h in cells]
            continue
        if not any(cells):
            continue
        rows.append({headers[j]: cells[j] for j in range(min(len(headers), len(cells)))})
    return rows


def _read_upload(file: UploadFile) -> list[dict]:
    filename = (file.filename or "").lower()
    content = file.file.read()
    if filename.endswith(".xlsx"):
        return _read_xlsx(content)
    if filename.endswith(".csv"):
        return _read_csv(content)
    raise HTTPException(status_code=400, detail="Format non supporté (utilisez .csv ou .xlsx)")


def _read_grid(file: UploadFile) -> list[list[str]]:
    """Lit le fichier comme une grille brute de cellules (pour les formats à en-tête)."""
    filename = (file.filename or "").lower()
    content = file.file.read()
    if filename.endswith(".xlsx"):
        wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            return []
        return [
            [("" if v is None else str(v)).strip() for v in row]
            for row in ws.iter_rows(values_only=True)
        ]
    if filename.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="replace")
        return [[c.strip() for c in row] for row in csv.reader(io.StringIO(text))]
    raise HTTPException(status_code=400, detail="Format non supporté (utilisez .csv ou .xlsx)")


# ─── Endpoints d'import ───────────────────────────────────────────────────────


@router.post("/enseignants", response_model=ImportResult)
def import_enseignants(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _=Depends(require_admin),
) -> ImportResult:
    """Colonnes attendues : nom_complet, grade."""
    rows = _read_upload(file)
    created = updated = skipped = 0
    errors: list[str] = []

    for i, row in enumerate(rows, start=2):
        nom = row.get("nom_complet") or row.get("nom") or ""
        grade = row.get("grade") or ""
        if not nom or not grade:
            errors.append(f"Ligne {i}: nom_complet et grade requis")
            skipped += 1
            continue

        existing = db.query(Enseignant).filter(Enseignant.nom_complet == nom).first()
        if existing:
            existing.grade = grade
            updated += 1
        else:
            db.add(Enseignant(nom_complet=nom, grade=grade))
            created += 1

    db.commit()
    return ImportResult(created=created, updated=updated, skipped=skipped, errors=errors)


@router.post("/academic", response_model=ImportResult)
def import_academic(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _=Depends(require_admin),
) -> ImportResult:
    """
    Import des ECUEs d'UNE classe et UN semestre (un fichier = 1 classe + 1 semestre).

    Format attendu :
        - Un en-tête en haut donne le contexte (une valeur par ligne) :
            niveau    | L1
            classe    | Informatique
            semestre  | Semestre 1
        - Puis une ligne d'en-tête de tableau :
            ecue | heures_cm | heures_td | heures_tp
        - Puis une ligne par ECUE :
            Algèbre 1 | 20 | 10 | 5
            Analyse 1 | 15 |  8 | 12

    Import ADDITIF : les niveaux/classes existants sont réutilisés, les ECUEs
    déjà présentes (même intitulé, même classe+semestre) voient leurs heures
    mises à jour, les nouvelles sont ajoutées. Rien n'est supprimé.
    """
    grid = _read_grid(file)
    created = updated = skipped = 0
    errors: list[str] = []

    # 1) Lecture de l'en-tête méta (niveau / classe / semestre) et repérage de
    #    la ligne d'en-tête du tableau d'ECUEs.
    meta = {"niveau": "", "classe": "", "semestre": ""}
    table_header_idx = -1
    for idx, raw in enumerate(grid):
        cells = list(raw)
        if not any(c.strip() for c in cells):
            continue
        key = cells[0].strip().lower() if cells else ""
        # Ligne d'en-tête du tableau d'ECUEs ?
        if key in ("ecue", "intitule", "intitulé"):
            table_header_idx = idx
            break
        # Ligne méta « clé | valeur »
        if key in meta and len(cells) >= 2:
            meta[key] = cells[1].strip()

    if table_header_idx < 0:
        raise HTTPException(
            status_code=400,
            detail="Format invalide : ligne d'en-tête « ecue | heures_cm | heures_td | heures_tp » introuvable.",
        )

    niveau_nom = meta["niveau"]
    classe_nom = meta["classe"]
    semestre_nom = meta["semestre"]
    if not all([niveau_nom, classe_nom, semestre_nom]):
        raise HTTPException(
            status_code=400,
            detail="En-tête incomplet : niveau, classe et semestre sont requis en haut du fichier.",
        )

    # 2) Récupère / crée niveau → classe → semestre (additif).
    niveau = db.query(Niveau).filter(Niveau.nom == niveau_nom).first()
    if not niveau:
        niveau = Niveau(nom=niveau_nom)
        db.add(niveau)
        db.flush()

    classe = (
        db.query(Classe)
        .filter(Classe.niveau_id == niveau.id, Classe.nom == classe_nom)
        .first()
    )
    if not classe:
        classe = Classe(nom=classe_nom, niveau_id=niveau.id)
        db.add(classe)
        db.flush()

    semestre = db.query(Semestre).filter(Semestre.nom == semestre_nom).first()
    if not semestre:
        semestre = Semestre(nom=semestre_nom)
        db.add(semestre)
        db.flush()

    # 3) Colonnes du tableau d'ECUEs (positions par nom d'en-tête).
    table_headers = [c.strip().lower() for c in grid[table_header_idx]]

    def col(*names: str) -> int:
        for n in names:
            if n in table_headers:
                return table_headers.index(n)
        return -1

    i_ecue = col("ecue", "intitule", "intitulé")
    i_cm = col("heures_cm", "cm")
    i_td = col("heures_td", "td")
    i_tp = col("heures_tp", "tp")

    def cell_int(cells: list[str], pos: int) -> int:
        if pos < 0 or pos >= len(cells):
            return 0
        try:
            return max(0, min(99, int(float(cells[pos] or 0))))
        except (ValueError, TypeError):
            return 0

    # 4) Lignes d'ECUEs.
    for n, raw in enumerate(grid[table_header_idx + 1 :], start=table_header_idx + 2):
        cells = list(raw)
        if not any(c.strip() for c in cells):
            continue
        ecue_nom = cells[i_ecue].strip() if 0 <= i_ecue < len(cells) else ""
        if not ecue_nom:
            errors.append(f"Ligne {n}: intitulé d'ECUE manquant")
            skipped += 1
            continue

        heures_cm = cell_int(cells, i_cm)
        heures_td = cell_int(cells, i_td)
        heures_tp = cell_int(cells, i_tp)

        existing_ecue = (
            db.query(Ecue)
            .filter(
                Ecue.intitule == ecue_nom,
                Ecue.classe_id == classe.id,
                Ecue.semestre_id == semestre.id,
            )
            .first()
        )
        if existing_ecue:
            existing_ecue.heures_cm_defaut = heures_cm
            existing_ecue.heures_td_defaut = heures_td
            existing_ecue.heures_tp_defaut = heures_tp
            updated += 1
        else:
            db.add(
                Ecue(
                    intitule=ecue_nom,
                    classe_id=classe.id,
                    semestre_id=semestre.id,
                    heures_cm_defaut=heures_cm,
                    heures_td_defaut=heures_td,
                    heures_tp_defaut=heures_tp,
                )
            )
            created += 1

    db.commit()
    return ImportResult(created=created, updated=updated, skipped=skipped, errors=errors)


# ─── Téléchargement de modèles ────────────────────────────────────────────────


# Chaque modèle est défini comme une simple grille de lignes (« rows ») : la
# première représentation est universelle et fonctionne aussi bien pour un
# tableau plat (enseignants) que pour le format à en-tête (academic).
TEMPLATES: dict[str, dict] = {
    "enseignants": {
        "rows": [
            ["nom_complet", "grade"],
            ["Dr CHARFADINE MAHAMAT", "Maître Assistant"],
            ["Dr HAGGAR BACHAR SALIM", "Professeur"],
            ["MOUKHTAR HASSAN MAHAMAT", "Maître de Conférences"],
        ],
        # Largeurs de colonnes (XLSX).
        "widths": [32, 24],
    },
    "academic": {
        # Un fichier = 1 niveau + 1 classe + 1 semestre.
        #   En-tête méta « clé | valeur » en haut, puis le tableau d'ECUEs.
        "rows": [
            ["niveau", "L1"],
            ["classe", "Informatique"],
            ["semestre", "Semestre 1"],
            [],  # ligne vide de séparation
            ["ecue", "heures_cm", "heures_td", "heures_tp"],
            ["Algèbre 1", "20", "10", "5"],
            ["Analyse 1", "15", "8", "12"],
            ["Initiation à la programmation", "14", "8", "6"],
            ["Architecture des ordinateurs", "18", "6", "10"],
        ],
        "widths": [34, 12, 12, 12],
    },
}


def _csv_template(template: dict) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    for row in template["rows"]:
        writer.writerow(row)
    return buf.getvalue().encode("utf-8")


def _xlsx_template(template: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Modèle"
    for row in template["rows"]:
        ws.append(row)
    # Largeur des colonnes.
    for col_idx, width in enumerate(template.get("widths", []), start=1):
        ws.column_dimensions[chr(64 + col_idx)].width = width
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@router.get("/template/{kind}.{fmt}")
def download_template(
    kind: str,
    fmt: str,
    _=Depends(require_admin),
) -> StreamingResponse:
    """Télécharge un modèle CSV ou XLSX (kind: enseignants | academic ; fmt: csv | xlsx)."""
    if kind not in TEMPLATES:
        raise HTTPException(status_code=404, detail="Type de modèle inconnu")
    template = TEMPLATES[kind]
    if fmt == "csv":
        content = _csv_template(template)
        media = "text/csv"
        filename = f"modele_{kind}.csv"
    elif fmt == "xlsx":
        content = _xlsx_template(template)
        media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = f"modele_{kind}.xlsx"
    else:
        raise HTTPException(status_code=400, detail="Format invalide (csv ou xlsx)")

    return StreamingResponse(
        io.BytesIO(content),
        media_type=media,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
