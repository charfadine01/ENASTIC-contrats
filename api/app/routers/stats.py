"""
Statistiques : récapitulatif des heures et contrats par enseignant.

Les heures détaillées vivent dans Contract.contract_metadata["ecues"]. On les
agrège ici par enseignant (et par année académique si filtre), avec un export
Excel téléchargeable.
"""

import io

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import Contract, User
from app.security import require_admin

router = APIRouter(prefix="/stats", tags=["stats"])


def _aggregate(db: Session, annee: str | None) -> list[dict]:
    """Agrège par enseignant : nb de contrats + heures CM/TD/TP."""
    query = db.query(Contract)
    if annee:
        query = query.filter(Contract.academic_year == annee)

    par_ens: dict[str, dict] = {}
    for c in query.all():
        key = c.teacher_name
        row = par_ens.setdefault(
            key,
            {
                "enseignant": c.teacher_name,
                "grade": c.teacher_grade,
                "contrats": 0,
                "heures_cm": 0,
                "heures_td": 0,
                "heures_tp": 0,
            },
        )
        row["contrats"] += 1
        # Grade le plus récent (les contrats sont parcourus dans l'ordre DB).
        row["grade"] = c.teacher_grade
        ecues = (c.contract_metadata or {}).get("ecues", []) or []
        for e in ecues:
            row["heures_cm"] += int(e.get("heures_cm", 0) or 0)
            row["heures_td"] += int(e.get("heures_td", 0) or 0)
            row["heures_tp"] += int(e.get("heures_tp", 0) or 0)

    rows = list(par_ens.values())
    for r in rows:
        r["heures_total"] = r["heures_cm"] + r["heures_td"] + r["heures_tp"]
    # Tri par total d'heures décroissant, puis par nom.
    rows.sort(key=lambda r: (-r["heures_total"], r["enseignant"].lower()))
    return rows


@router.get("/enseignants")
def stats_enseignants(
    annee: str | None = Query(default=None, description="Filtre année académique (AAAA-AAAA)"),
    db: Session = Depends(get_db),
    _admin: User = Depends(require_admin),
) -> dict:
    """Récapitulatif par enseignant (JSON) + liste des années disponibles."""
    rows = _aggregate(db, annee)
    annees = sorted(
        {c.academic_year for c in db.query(Contract.academic_year).distinct()},
        reverse=True,
    )
    totaux = {
        "enseignants": len(rows),
        "contrats": sum(r["contrats"] for r in rows),
        "heures_cm": sum(r["heures_cm"] for r in rows),
        "heures_td": sum(r["heures_td"] for r in rows),
        "heures_tp": sum(r["heures_tp"] for r in rows),
    }
    totaux["heures_total"] = totaux["heures_cm"] + totaux["heures_td"] + totaux["heures_tp"]
    return {"annee": annee, "annees": annees, "lignes": rows, "totaux": totaux}


@router.get("/enseignants/export")
def stats_enseignants_export(
    annee: str | None = Query(default=None),
    db: Session = Depends(get_db),
    _admin: User = Depends(require_admin),
) -> StreamingResponse:
    """Export Excel du récapitulatif par enseignant."""
    rows = _aggregate(db, annee)

    wb = Workbook()
    ws = wb.active
    ws.title = "Heures par enseignant"

    titre = f"Récapitulatif des heures par enseignant" + (f" — {annee}" if annee else " — toutes années")
    ws.append([titre])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([])

    headers = ["Enseignant", "Grade", "Contrats", "Heures CM", "Heures TD", "Heures TP", "Total heures"]
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    for r in rows:
        ws.append([
            r["enseignant"],
            r["grade"],
            r["contrats"],
            r["heures_cm"],
            r["heures_td"],
            r["heures_tp"],
            r["heures_total"],
        ])

    # Ligne de total général.
    if rows:
        ws.append([])
        ws.append([
            "TOTAL",
            "",
            sum(r["contrats"] for r in rows),
            sum(r["heures_cm"] for r in rows),
            sum(r["heures_td"] for r in rows),
            sum(r["heures_tp"] for r in rows),
            sum(r["heures_total"] for r in rows),
        ])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)

    widths = [34, 22, 10, 11, 11, 11, 13]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    suffix = (annee or "toutes-annees").replace("/", "-")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="heures-enseignants-{suffix}.xlsx"'
        },
    )
